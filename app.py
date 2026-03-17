from flask import Flask, render_template_string, request, send_file
import pandas as pd
import json
import io
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)

CAMPOS_PRODUCTO = ['numItem', 'cantidad', 'descripcion', 'precioUni', 'montoDescu', 'ventaGravada']

# Estilo "Verdecito" restaurado
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Convertidor DE JSON</title>
    <style>
        body { font-family: sans-serif; display: flex; justify-content: center; align-items: center; height: 100vh; margin: 0; background: #e8f5e9; }
        .card { background: white; padding: 40px; border-radius: 10px; box-shadow: 0 4px 12px rgba(0,0,0,0.1); text-align: center; border-top: 8px solid #4caf50; }
        h2 { color: #2e7d32; }
        button { background: #4caf50; color: white; border: none; padding: 12px 24px; cursor: pointer; border-radius: 4px; font-size: 16px; font-weight: bold; }
        button:hover { background: #388e3c; }
        p { color: #666; }
    </style>
</head>
<body>
    <div class="card">
        <h2>📂 Extractor de JSON</h2>
        <p>Extractor  de Json </p>
        <form method="post" enctype="multipart/form-data">
            <input type="file" name="file" accept=".json" multiple required><br>
            <button type="submit">Generar Reporte Excel</button>
        </form>
    </div>
</body>
</html>
'''

def estilizar_excel(ws, titulo):
    # Centrar y dar formato al título en la fila 1
    ws.merge_cells('A1:B1')
    top_cell = ws['A1']
    top_cell.value = titulo.upper()
    top_cell.font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    top_cell.alignment = Alignment(horizontal='center')
    # Color verde oscuro para el encabezado del Excel también
    top_cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    
    # Ajuste de anchos para que todo quepa bien
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 60

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        files = request.files.getlist('file')
        if files:
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                for idx, file in enumerate(files):
                    try:
                        data = json.load(file)
                        
                        # Identificación del tipo de documento
                        tipo = data.get('identificacion', {}).get('tipoDte', '01')
                        nombre_doc = "COMPROBANTE DE CRÉDITO FISCAL" if tipo == '03' else "FACTURA CONSUMIDOR FINAL"
                        
                        # Datos que marcaste en amarillo (Emisor, Receptor, MH)
                        secciones_interes = {
                            '1. IDENTIFICACION': data.get('identificacion', {}),
                            '2. EMISOR': data.get('emisor', {}),
                            '3. RECEPTOR': data.get('receptor', {}),
                            '4. RESPUESTA MH (SELLOS)': data.get('responseMH', {})
                        }
                        df_general = pd.json_normalize(secciones_interes, sep='.').T.reset_index()
                        df_general.columns = ['Campo', 'Valor']

                        # Cuerpo del documento (Productos)
                        df_prod = pd.DataFrame(data.get('cuerpoDocumento', []))
                        if not df_prod.empty:
                            df_prod = df_prod[[c for c in CAMPOS_PRODUCTO if c in df_prod.columns]]

                        # Resumen de totales
                        df_res = pd.json_normalize(data.get('resumen', {}), sep='.').T.reset_index()
                        df_res.columns = ['Campo', 'Valor']

                        # Crear la hoja
                        sheet_name = f"Documento_{idx+1}"
                        df_general.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)
                        
                        ws = writer.sheets[sheet_name]
                        estilizar_excel(ws, nombre_doc)

                        # Escribir productos debajo de los datos generales
                        gap = 4 + len(df_general)
                        ws.cell(row=gap, column=1, value="--- DETALLE DE PRODUCTOS ---").font = Font(bold=True)
                        df_prod.to_excel(writer, index=False, sheet_name=sheet_name, startrow=gap)

                        # Escribir totales al final
                        gap_res = gap + len(df_prod) + 2
                        ws.cell(row=gap_res, column=1, value="--- TOTALES Y RESUMEN ---").font = Font(bold=True)
                        df_res.to_excel(writer, index=False, header=False, sheet_name=sheet_name, startrow=gap_res)

                    except Exception: continue

            output.seek(0)
            return send_file(output, download_name="Reporte de Json .xlsx", as_attachment=True)
            
    return render_template_string(HTML_TEMPLATE)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)